from openpyxl import load_workbook

wb = load_workbook('excel.xlsx')
ws = wb.active


def get_col_value():
    items = []
    for i in range(1, ws.max_row+1):
        col_value = str(ws.cell(row=i, column=4).value)
        items.append(col_value)

    return items
